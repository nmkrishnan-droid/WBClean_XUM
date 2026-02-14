import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from collections import Counter
import requests
import win32com.client as win32
import os


class WBClean_XUM:
    def XUM_TransposeSheet(self, srcPath,destinationSheetName, srcSheetName=None):
        wb = load_workbook(srcPath, data_only=False)
        ws = wb[srcSheetName] if srcSheetName else wb.worksheets[0]


        min_row, max_row = ws.min_row, ws.max_row
        min_col, max_col = ws.min_column, ws.max_column

        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = destinationSheetName or (srcSheetName + "_T")


        grid = [
            [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)]
            for r in range(min_row, max_row + 1)
        ]

        tgrid = list(zip(*grid))  # tuples
        for r, row_vals in enumerate(tgrid, start=1):
            for c, v in enumerate(row_vals, start=1):
                out_ws.cell(row=r, column=c, value=v)

        tableList = []
        for row in out_ws.iter_rows():
            rowList = []
            for cell in row:
                rowList.append(cell.value)
            tableList.append(rowList)

        return tableList

    def XUM_LLMFormat(self, prompt_ReqFeildString, prompt_ReqJSONOutputString, prompt_SampleData, groqModel, Key,
                      APIUrl,
                      temperature=0, maxTokens=512, contentType="application/json", prompt_FullCustom=False,
                      prompt_Full=None):
        prompt = f"""
        You are a data-mapping assistant.

        The SAMPLE DATA is given in /{{Key:value}} format where Key is the column Name and the value is sample values for that column

        Map the SAMPLE DATA KEYS to the REQUIRED fields below.

        REQUIRED FIELDS:
        {prompt_ReqFeildString}

        SAMPLE DATA:
        {prompt_SampleData}

        RULES:
        - Choose best column if multiple match
        - Use null if none apply
        - alt_cols must be a JSON array
        - RETURN ONLY VALID JSON. NO TEXT.

        EXPECTED JSON FORMAT:
        {prompt_ReqJSONOutputString}

        Return ONLY a JSON object. Do not use markdown, do not use code fences, do not add commentary.

        """

        if prompt_FullCustom:
            prompt = prompt_Full

        payload = {
            "model": groqModel,
            "messages": [
                {"role": "user", "content": prompt}
            ],
            "temperature": temperature,
            "max_tokens": maxTokens
        }

        r = requests.post(
            APIUrl,
            headers={
                "Authorization": f"Bearer {Key}",
                "Content-Type": contentType,
            },
            json=payload,
            timeout=60,
        )

        if r.status_code != 200:
            raise RuntimeError(
                f"GROQ ERROR {r.status_code}: {r.text}"
            )

        content = r.json()["choices"][0]["message"]["content"]
        try:
            data = content
        except Exception as e:
            raise RuntimeError(f"LLM did not return valid JSON. Raw content:\n{content}") from e

        return data

    def XUM_DeleteRows(self, Table2dArray, rowIndexList):
        rows_to_drop = set(rowIndexList)  # 0-based indices

        newTable = [
            row for idx, row in enumerate(Table2dArray)
            if idx not in rows_to_drop
        ]
        return newTable

    def XUM_DeleteColumns(self, Table2dArray, colIndexList):
        cols_to_drop = set(colIndexList)

        newTable = []
        for row in Table2dArray:
            # keep values whose 1-based index is NOT in cols_to_drop
            newRow = [val for idx, val in enumerate(row, start=1) if idx not in cols_to_drop]
            newTable.append(newRow)

        return newTable

    def XUM_TextPresenceRegex(self, x, pattern):
        if x is None:
            return False
        return pattern.search(str(x)) is not None

    def XUM_XLSConversion(self, xlsPath):
        xls_path = os.path.abspath(xlsPath)
        if xlsPath is None:
            root, _ = os.path.splitext(xls_path)
            xlsx_path = root + ".xlsx"
        xlsxPath = os.path.abspath(xlsPath)

        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(xls_path)
        wb.SaveAs(xlsxPath, FileFormat=51)  # 51 = xlOpenXMLWorkbook (.xlsx)
        wb.Close(False)
        excel.Quit()

        return xlsxPath

    def XUM_Clean(self, filePath, pattern, returnDF=True, remove_none=True, getImpFeatures=False,
                  prompt_ReqFeildString=None, prompt_ReqJSONOutputString=None, Key=None,
                  sheetName=None,
                  destinationSheet="WBClean_XUM",
                  APIUrl="https://api.groq.com/openai/v1/chat/completions", groqModel="llama-3.3-70b-versatile",
                  contentType="application/json", temperature=0, maxTokens=512):
        transposedList = self.XUM_TransposeSheet(
            src_path=filePath,
            src_sheet_name=sheetName,
            dst_sheet_name=destinationSheet
        )

        out = []
        foundIndex = []

        for i in transposedList:
            for ind, value in enumerate(i, start=1):
                if value == None:
                    out.append(ind)
                elif self.XUM_TextPresenceRegex(value, re.compile(rf'\b(?:{pattern})\b', re.I)):
                    foundIndex.append(ind)
                    break


        captureCols = sorted(x for x in out if x < foundIndex[-1])

        freq = Counter(captureCols)
        result = sorted(
            [(num, cnt) for num, cnt in freq.items() if cnt > 1],
            key=lambda x: x[1],
            reverse=True
        )
        nums_only = [num for num, cnt in result]

        newTable = self.XUM_DeleteColumns(Table2dArray=transposedList, colIndexList=nums_only)


        headerColumns = [x for i in newTable for ind, x in enumerate(i) if ind == 0]
        row1Values = [x for i in newTable for ind, x in enumerate(i) if ind == 1]

        if getImpFeatures:
            testSample = dict(zip(headerColumns, row1Values))
            columnMapping = self.XUM_LLMFormat(
                prompt_ReqFeildString=prompt_ReqFeildString,
                prompt_ReqJSONOutputString=prompt_ReqJSONOutputString,
                prompt_SampleData=testSample,
                Key=Key,
                APIUrl=APIUrl,
                groqModel=groqModel,
                temperature=temperature,
                maxTokens=maxTokens,
                contentType=contentType
            )

            rowsWeDontNeed = []
            for rowInd, row in enumerate(newTable):
                if row[0] == None:
                    rowsWeDontNeed.append(rowInd)
                elif row[0] in columnMapping:
                    pass
                else:
                    rowsWeDontNeed.append(rowInd)


            TableWeNeed = self.XUM_DeleteRows(Table2dArray=newTable, rowIndexList=rowsWeDontNeed)

            newTable = TableWeNeed

        else:
            newTable = newTable

        transposeToNormal = [list(r) for r in zip(*newTable)]


        if remove_none:
            rowsToDelete = [idx for idx, row in enumerate(transposeToNormal) if all(v is None for v in row)]
            table = self.XUM_DeleteRows(Table2dArray=transposeToNormal, rowIndexList=rowsToDelete)
        else:
            table = transposeToNormal

        df = pd.DataFrame(table[1:], columns=table[0])

        return df if returnDF else df.to_excel(f"{destinationSheet}.xlsx")


